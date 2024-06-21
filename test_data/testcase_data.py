import openpyxl


class Data:

    @staticmethod
    def getTestData(test_case_name, file_path):
        lst = []
        book = openpyxl.load_workbook(file_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            Dict = {}
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                lst.append(Dict)

        return lst
